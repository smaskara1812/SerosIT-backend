"""Dumps every table in the app into one .xlsx workbook, one sheet per table.

Just a convenience for eyeballing/sharing the whole dev DB at once — not part
of the production migration path (see build_migration_map for that).

    python manage.py export_all_tables
    python manage.py export_all_tables --output ~/Desktop/serosit_dump.xlsx
    python manage.py export_all_tables --tables mst_rig,mst_company
"""

import datetime
import os

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Excel hard-limits sheet names to 31 chars and disallows []:*?/\\.
_INVALID_SHEET_CHARS = set('[]:*?/\\')


def _sheet_name(table, used):
    name = "".join(c for c in table if c not in _INVALID_SHEET_CHARS)[:31]
    if name not in used:
        used.add(name)
        return name
    # Collision after truncation (rare) — suffix with a counter, still <=31.
    for i in range(2, 1000):
        suffix = f"~{i}"
        candidate = name[: 31 - len(suffix)] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise CommandError(f"Could not find a unique sheet name for table '{table}'")


def _cell_value(value):
    # openpyxl refuses tz-aware datetimes outright (Excel has no timezone
    # concept) — strip tzinfo rather than converting, so the wall-clock time
    # shown matches what the app displays (already localized on read).
    if isinstance(value, datetime.datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    # Everything else openpyxl writes natively (date/int/float/str/bool/
    # None) passes through; anything unexpected (Decimal, etc.) is
    # stringified so the write never raises on an unrecognized type.
    if value is None or isinstance(value, (str, int, float, bool, datetime.date, datetime.datetime)):
        return value
    return str(value)


class Command(BaseCommand):
    help = "Export every model's table to a single .xlsx file, one sheet per table."

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            default=None,
            help="Output .xlsx path (default: docs/full_export_<timestamp>.xlsx)",
        )
        parser.add_argument(
            "--tables",
            default=None,
            help="Comma-separated db_table names to export (default: every table in the app)",
        )

    def handle(self, *args, **options):
        models = sorted(
            apps.get_app_config("core").get_models(), key=lambda m: m._meta.db_table
        )

        if options["tables"]:
            wanted = {t.strip() for t in options["tables"].split(",") if t.strip()}
            models = [m for m in models if m._meta.db_table in wanted]
            missing = wanted - {m._meta.db_table for m in models}
            if missing:
                raise CommandError(f"Unknown table(s): {', '.join(sorted(missing))}")

        output = options["output"]
        if not output:
            from django.conf import settings

            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output = os.path.join(settings.BASE_DIR, "docs", f"full_export_{stamp}.xlsx")

        wb = Workbook()
        wb.remove(wb.active)  # drop the default blank sheet
        used_names = set()
        header_font = Font(bold=True)

        for model in models:
            table = model._meta.db_table
            fields = model._meta.fields
            # attname (not column) is what .values() needs to select a FK
            # by its id without a join — for most fields the two match, but
            # a few (e.g. ItAssetHolder.employee -> db_column="emp_id")
            # give the field a db_column that differs from Django's own
            # <field>_id attname.
            columns = [f.attname for f in fields]

            ws = wb.create_sheet(_sheet_name(table, used_names))
            ws.append(columns)
            for cell in ws[1]:
                cell.font = header_font
            ws.freeze_panes = "A2"

            row_count = 0
            for row in model.objects.order_by("pk").values(*columns).iterator(chunk_size=2000):
                ws.append([_cell_value(row[c]) for c in columns])
                row_count += 1

            # Rough auto-fit: width = longest of header/seen values, capped
            # so one huge free-text column doesn't blow out the sheet.
            for i, col in enumerate(columns, start=1):
                header_len = len(col)
                ws.column_dimensions[get_column_letter(i)].width = min(max(header_len + 2, 10), 40)

            self.stdout.write(f"  {table}: {row_count} rows")

        os.makedirs(os.path.dirname(output), exist_ok=True)
        wb.save(output)
        self.stdout.write(self.style.SUCCESS(f"\nWrote {len(models)} sheets to {output}"))

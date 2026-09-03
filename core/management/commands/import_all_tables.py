"""Reloads every sheet in an .xlsx (as produced by export_all_tables) back
into its table — the companion "restore" half of that command.

Writes go through raw parameterized INSERTs, not the Django ORM's
save()/bulk_create() — two real landmines forced that:
  1. bulk_create() calls each field's pre_save(), which unconditionally
     overwrites auto_now/auto_now_add columns (cr_dt/mod_dt-style audit
     timestamps) with the current time, destroying the imported value.
  2. Django's AutoField refuses an explicit 0 as a primary key outright
     (`ValueError: The database backend does not accept 0 as a value for
     AutoField`) — and this dev DB genuinely has PK 0 rows (e.g.
     mst_user.user_id=0, a legacy sentinel).
Raw SQL sidesteps both: it writes exactly the literal value that was in the
sheet, with no per-field ORM side effects.

Works against MySQL (dev) and SQL Server (prod) — both backends implement
Django's standard `connection.constraint_checks_disabled()` (MySQL via
`SET FOREIGN_KEY_CHECKS`, mssql-django via `ALTER TABLE ... NOCHECK
CONSTRAINT ALL` on every table), so this writes tables in whatever order the
sheets happen to be in rather than hand-computing FK insertion order — safe
because the export is already a self-consistent snapshot of a live DB, so
there's nothing invalid to protect against while checks are off.
mssql-django's cursor also transparently rewrites this file's `%s`
placeholders to pyodbc's `?`, so the parameterized SQL itself needs no
vendor branching — only identifier quoting (`` ` `` vs `[ ]`) and SQL
Server's `IDENTITY_INSERT` (required for *any* explicit identity-column
value, not just 0 — MySQL only needs the explicit-0 workaround) differ.

    python manage.py import_all_tables docs/full_export_20260901.xlsx --dry-run
    python manage.py import_all_tables docs/full_export_20260901.xlsx --truncate
    python manage.py import_all_tables docs/full_export_20260901.xlsx --tables mst_rig,mst_company
"""

import os
from decimal import Decimal

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction
from django.db.models import DecimalField
from django.db.models.fields import AutoFieldMixin
from openpyxl import load_workbook

SUPPORTED_VENDORS = {"mysql", "microsoft"}


def _quote(name, vendor):
    return f"[{name}]" if vendor == "microsoft" else f"`{name}`"


def _has_identity_pk(model, db_columns):
    """True if this table's PK is an identity/AUTO_INCREMENT column *and*
    the sheet actually carries a value for it — SQL Server's IDENTITY_INSERT
    requires every row to supply the identity column explicitly, so it's
    only safe to turn on when that column is really present."""
    pk = model._meta.pk
    return isinstance(pk, AutoFieldMixin) and pk.column in db_columns


class Command(BaseCommand):
    help = "Import every matching sheet in an .xlsx back into its table (companion to export_all_tables)."

    def add_arguments(self, parser):
        parser.add_argument("input", help="Path to the .xlsx file to import")
        parser.add_argument(
            "--tables",
            default=None,
            help="Comma-separated db_table names to import (default: every table with a matching sheet)",
        )
        parser.add_argument(
            "--truncate",
            action="store_true",
            help="Empty each targeted table (TRUNCATE) before importing its sheet",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Only parse and validate the sheets — no database writes at all",
        )

    def handle(self, *args, **options):
        path = options["input"]
        if not os.path.isfile(path):
            raise CommandError(f"File not found: {path}")
        vendor = connection.vendor
        if vendor not in SUPPORTED_VENDORS:
            raise CommandError(
                f"This command supports MySQL and SQL Server only — the current connection is "
                f"'{vendor}'."
            )

        wb = load_workbook(path, read_only=True, data_only=True)

        models = list(apps.get_app_config("core").get_models())
        by_table = {m._meta.db_table: m for m in models}

        # export_all_tables truncates a table name to Excel's 31-char sheet
        # limit — a straight truncated match is enough since no two tables
        # in this app collide once cut to 31 chars.
        sheet_for_table = {t: t[:31] for t in by_table if t[:31] in wb.sheetnames}

        if options["tables"]:
            wanted = {t.strip() for t in options["tables"].split(",") if t.strip()}
            unknown = wanted - set(by_table)
            if unknown:
                raise CommandError(f"Unknown table(s): {', '.join(sorted(unknown))}")
            missing = wanted - set(sheet_for_table)
            if missing:
                raise CommandError(f"No sheet found in this file for: {', '.join(sorted(missing))}")
            sheet_for_table = {t: s for t, s in sheet_for_table.items() if t in wanted}
        else:
            skipped = set(by_table) - set(sheet_for_table)
            if skipped:
                self.stdout.write(
                    self.style.WARNING(
                        f"No sheet found for {len(skipped)} table(s), skipping: {', '.join(sorted(skipped))}"
                    )
                )

        # table -> (db column list, [row tuple, ...])
        planned = {}
        total_rows = 0

        for table, sheet_name in sheet_for_table.items():
            model = by_table[table]
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                planned[table] = ([], [])
                continue
            header = list(header)
            fields_by_attname = {f.attname: f for f in model._meta.fields}
            unknown_cols = [h for h in header if h not in fields_by_attname]
            if unknown_cols:
                raise CommandError(
                    f"{table}: sheet has column(s) {unknown_cols} the current model doesn't have — "
                    f"the schema changed since this file was exported."
                )
            missing_cols = [a for a in fields_by_attname if a not in header]
            if missing_cols:
                self.stdout.write(
                    self.style.WARNING(
                        f"  {table}: sheet is missing column(s) {missing_cols} — those columns will "
                        f"get the database's own column default on every imported row."
                    )
                )

            db_columns = [fields_by_attname[a].column for a in header]
            fields = [fields_by_attname[a] for a in header]

            row_tuples = []
            for values in rows_iter:
                if all(v is None for v in values):
                    continue
                out = []
                for field, raw in zip(fields, values):
                    value = raw
                    if raw is None and not field.null and getattr(field, "empty_strings_allowed", False):
                        # Excel can't distinguish an empty string from a
                        # blank cell — openpyxl reads both back as None.
                        # A NOT NULL string column's real empty value was
                        # '', not NULL (e.g. mst_cost_centre_type had a
                        # genuine '' row that a bare None would reject).
                        value = ""
                    elif isinstance(field, DecimalField) and isinstance(raw, float):
                        # Decimal(float) directly introduces binary-float
                        # noise (e.g. 12.10 -> 12.09999999999999964...) —
                        # round-tripping through str keeps the exact value.
                        value = Decimal(str(raw))
                    out.append(value)
                row_tuples.append(tuple(out))

            planned[table] = (db_columns, row_tuples)
            total_rows += len(row_tuples)
            self.stdout.write(f"  {table}: {len(row_tuples)} row(s) parsed")

        if options["dry_run"]:
            self.stdout.write(
                self.style.WARNING(
                    f"\nDry run: {total_rows} row(s) across {len(sheet_for_table)} table(s) parsed "
                    f"successfully. Nothing was written."
                )
            )
            return

        with transaction.atomic():
            with connection.constraint_checks_disabled():
                if vendor == "mysql":
                    with connection.cursor() as cur:
                        # Without this, MySQL treats an explicit 0 in an
                        # AUTO_INCREMENT column as "assign the next value"
                        # rather than literally storing 0 — and this DB has
                        # a real PK-0 sentinel row (mst_user.user_id=0,
                        # "System"), which would otherwise silently collide
                        # with PK 1. SQL Server has no equivalent quirk —
                        # IDENTITY_INSERT below handles explicit values,
                        # 0 included, uniformly.
                        cur.execute("SET SESSION sql_mode = CONCAT(@@sql_mode, ',NO_AUTO_VALUE_ON_ZERO')")

                if options["truncate"]:
                    self.stdout.write("Truncating targeted tables...")
                    with connection.cursor() as cur:
                        for table in sheet_for_table:
                            cur.execute(f"TRUNCATE TABLE {_quote(table, vendor)}")

                for table, (db_columns, row_tuples) in planned.items():
                    if not row_tuples:
                        continue
                    # Printed before, not just after, the insert — a big
                    # table's executemany() is one blocking call with no
                    # progress in between, so without this line the tool
                    # goes silent for however long that table takes and
                    # looks hung rather than merely slow.
                    self.stdout.write(f"  {table}: importing {len(row_tuples)} row(s)...")
                    self.stdout.flush()
                    model = by_table[table]
                    quoted_table = _quote(table, vendor)
                    col_list = ", ".join(_quote(c, vendor) for c in db_columns)
                    placeholders = ", ".join(["%s"] * len(db_columns))
                    sql = f"INSERT INTO {quoted_table} ({col_list}) VALUES ({placeholders})"
                    # SQL Server refuses to write an explicit value into an
                    # identity column at all unless IDENTITY_INSERT is ON
                    # for that table — and only one table may have it ON
                    # per session, so it's scoped tightly around this one
                    # table's insert rather than set once up front.
                    identity_wrap = vendor == "microsoft" and _has_identity_pk(model, db_columns)
                    with connection.cursor() as cur:
                        if vendor == "microsoft":
                            # Default pyodbc executemany() sends one round
                            # trip per row — for a table like mst_employee
                            # (tens of thousands of rows) that's the
                            # difference between seconds and many minutes,
                            # with zero output in between either way (this
                            # command only prints once a whole table is
                            # done), so it looks identical to "stuck". This
                            # is pyodbc's own documented batching flag;
                            # reaching for the raw cursor through Django's
                            # and mssql-django's wrapper layers because
                            # neither exposes it directly. Best-effort: if
                            # a future wrapper version changes shape, skip
                            # the speedup rather than fail the import over it.
                            try:
                                cur.cursor.cursor.fast_executemany = True
                            except AttributeError:
                                pass
                        if identity_wrap:
                            cur.execute(f"SET IDENTITY_INSERT {quoted_table} ON")
                        cur.executemany(sql, row_tuples)
                        if identity_wrap:
                            cur.execute(f"SET IDENTITY_INSERT {quoted_table} OFF")
                    self.stdout.write(f"  {table}: {len(row_tuples)} row(s) imported")

                if vendor == "mysql":
                    with connection.cursor() as cur:
                        cur.execute("SET SESSION sql_mode = @@GLOBAL.sql_mode")

        self.stdout.write(
            self.style.SUCCESS(f"\nImported {total_rows} row(s) across {len(sheet_for_table)} table(s) from {path}")
        )
